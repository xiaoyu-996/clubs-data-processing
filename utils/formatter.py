#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
格式化输出工具模块
提供Excel文件格式化、报告生成等功能
"""

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from typing import Dict, List, Any
import os


class ExcelFormatter:
    """Excel文件格式化处理类"""
    
    @staticmethod
    def auto_adjust_column_width(worksheet, max_width: int = 50):
        """
        自动调整列宽
        
        Args:
            worksheet: openpyxl工作表对象
            max_width: 最大列宽限制
        """
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def apply_header_style(worksheet, row: int = 1):
        """
        应用表头样式
        
        Args:
            worksheet: openpyxl工作表对象
            row: 表头行号
        """
        header_font = Font(name='宋体', size=11, bold=True)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
    
    @staticmethod
    def save_with_formatting(df: pd.DataFrame, output_file: str, 
                            sheet_name: str = 'Sheet1', 
                            auto_width: bool = True):
        """
        保存DataFrame到Excel并应用格式
        
        Args:
            df: 数据框
            output_file: 输出文件路径
            sheet_name: 工作表名称
            auto_width: 是否自动调整列宽
        """
        # 确保输出目录存在
        output_dir = os.path.dirname(output_file)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        # 保存到Excel
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            worksheet = writer.sheets[sheet_name]
            
            # 应用表头样式
            ExcelFormatter.apply_header_style(worksheet)
            
            # 自动调整列宽
            if auto_width:
                ExcelFormatter.auto_adjust_column_width(worksheet)


class ReportGenerator:
    """报告生成器类"""
    
    @staticmethod
    def generate_markdown_report(title: str, sections: List[Dict[str, Any]], 
                                 output_file: str):
        """
        生成Markdown格式报告
        
        Args:
            title: 报告标题
            sections: 报告内容段落列表
            output_file: 输出文件路径
        """
        lines = []
        
        # 标题
        lines.append(f"# {title}\n")
        lines.append(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        lines.append("---\n")
        
        # 各个段落
        for section in sections:
            section_title = section.get('title', '')
            section_content = section.get('content', '')
            section_type = section.get('type', 'text')
            
            if section_title:
                lines.append(f"\n## {section_title}\n")
            
            if section_type == 'list':
                for item in section_content:
                    lines.append(f"- {item}\n")
            elif section_type == 'table':
                # 简单表格
                if isinstance(section_content, list) and section_content:
                    # 表头
                    headers = section_content[0]
                    lines.append(f"| {' | '.join(str(h) for h in headers)} |\n")
                    lines.append(f"| {' | '.join(['---'] * len(headers))} |\n")
                    
                    # 数据行
                    for row in section_content[1:]:
                        lines.append(f"| {' | '.join(str(c) for c in row)} |\n")
            else:
                lines.append(f"{section_content}\n")
        
        # 写入文件
        with open(output_file, 'w', encoding='utf-8') as f:
            f.writelines(lines)
    
    @staticmethod
    def generate_statistics_report(stats: Dict[str, Any], output_file: str):
        """
        生成统计报告
        
        Args:
            stats: 统计数据字典
            output_file: 输出文件路径
        """
        sections = []
        
        # 总体统计
        if 'summary' in stats:
            summary_content = []
            for key, value in stats['summary'].items():
                summary_content.append(f"**{key}**: {value}")
            
            sections.append({
                'title': '总体统计',
                'content': summary_content,
                'type': 'list'
            })
        
        # 详细统计
        if 'details' in stats:
            for detail_title, detail_data in stats['details'].items():
                if isinstance(detail_data, dict):
                    detail_content = []
                    for k, v in detail_data.items():
                        detail_content.append(f"**{k}**: {v}")
                    
                    sections.append({
                        'title': detail_title,
                        'content': detail_content,
                        'type': 'list'
                    })
        
        ReportGenerator.generate_markdown_report(
            title=stats.get('title', '统计报告'),
            sections=sections,
            output_file=output_file
        )


class DataMerger:
    """数据合并工具类"""
    
    @staticmethod
    def merge_club_info(club_list: List[str]) -> str:
        """
        合并社团信息
        
        Args:
            club_list: 社团名称列表
            
        Returns:
            合并后的社团名称字符串（逗号分隔）
        """
        if not club_list:
            return ""
        
        all_clubs = []
        for clubs in club_list:
            if pd.notna(clubs) and clubs != '':
                # 按逗号分割
                club_names = [club.strip() for club in str(clubs).split(',')]
                all_clubs.extend(club_names)
        
        # 去重并保持顺序
        unique_clubs = []
        for club in all_clubs:
            if club and club not in unique_clubs:
                unique_clubs.append(club)
        
        return ', '.join(unique_clubs)
    
    @staticmethod
    def smart_merge_values(values: List[Any], field_type: str = 'general') -> str:
        """
        智能合并字段值
        
        Args:
            values: 值列表
            field_type: 字段类型（contact/qq/name/class/general）
            
        Returns:
            合并后的最佳值
        """
        from .normalizer import DataNormalizer
        
        if not values:
            return ""
        
        # 根据字段类型标准化
        if field_type == 'contact':
            normalized_values = [DataNormalizer.normalize_contact(v) for v in values]
        elif field_type == 'qq':
            normalized_values = [DataNormalizer.normalize_qq(v) for v in values]
        elif field_type == 'name':
            normalized_values = [DataNormalizer.normalize_name(v) for v in values]
        elif field_type == 'class':
            normalized_values = [DataNormalizer.normalize_class_info(v) for v in values]
        else:
            normalized_values = [str(v).strip() for v in values if pd.notna(v) and str(v).strip()]
        
        # 去除None值
        valid_values = [v for v in normalized_values if v is not None and v != '']
        
        if not valid_values:
            return ""
        
        # 去重
        unique_values = []
        for value in valid_values:
            if value not in unique_values:
                unique_values.append(value)
        
        if len(unique_values) == 1:
            return unique_values[0]
        
        # 对于联系方式和QQ号，选择最标准的
        if field_type in ['contact', 'qq']:
            best_value = unique_values[0]
            for value in unique_values:
                if field_type == 'contact' and len(value) == 11:
                    best_value = value
                    break
                elif field_type == 'qq' and 5 <= len(value) <= 11:
                    best_value = value
                    break
            return best_value
        
        # 对于班级，选择最完整的
        if field_type == 'class':
            return max(unique_values, key=len)
        
        return unique_values[0]
